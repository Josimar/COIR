import shutil
import Apoio.tickets
import pandas as pd
import math 
from icecream import ic
from os.path import isfile, join, basename, exists
from datetime import datetime

# ===================================================================================================
# Formata campos contendo strings para float
# ===================================================================================================
def sanitiza_moeda(moeda):
    #Se o valor é uma string então remove os simbolo $ e delimitadores
    #caso contrário, o valor é numérico e pode ser convertido
    if isinstance(moeda, str):
        return moeda.replace('CONTINUA...','0').replace('T - Liquidação pelo Bruto','0').replace('ON NM','0').replace('.','').replace(',','.').replace('R$','').replace('$','').replace('NM','0').replace('ON','0').replace('N1','0').replace('N2','0').replace('C O N T I N U A   ','0').replace("| D",'').replace("|D ",'').replace(" |D",'').replace('D','').replace("| C",'').replace("|C",'').replace("|",'').replace("| ",'').replace("0| ",'0').replace("0|",'0').replace('Compra Opções','0').replace('0 ay Trade (proj)','0').replace('+0 Custos Impostos','0')
    return moeda

# ===================================================================================================
# Sanitinizar Especificação do título
# ===================================================================================================
def sanitiza_especificacao_titulo(especificacao_titulo):
    especificacao_titulo = especificacao_titulo.str.replace(' NM','',regex=False).str.replace('  N1','',regex=False).str.replace(' N1','',regex=False).str.replace('N1','',regex=False).str.replace(' N2','',regex=False).str.replace('  N2','',regex=False).str.replace(' EDJ','',regex=False).str.replace(' EDB','',regex=False).str.replace(' ED','',regex=False).str.replace(' MA','',regex=False).str.replace(' M2 ','',regex=False).str.replace(' M2','',regex=False).str.replace(' MB','',regex=False).str.replace(' DR1','',regex=False).str.replace(' DR2','',regex=False).str.replace(' DR3','',regex=False).str.replace(' DRE','',regex=False).str.replace('DRN A','',regex=False).str.replace('DRN','',regex=False).str.replace(' EJS','',regex=False).str.replace(' EJ','',regex=False).str.replace(' EB','',regex=False).str.replace(' *','',regex=False)
    return especificacao_titulo

# ===================================================================================================
# Formata campos das Notas BMF com mais de uma página por nota
# ===================================================================================================
def sanitiza_nota_bmf(value):
    if isinstance(value, str):
        return value.replace('Venda disponível','0').replace('IRRF','0').replace('Outros','0').replace('Compra disponível','0').replace('IRRF Day Trade (proj.)','0').replace('0 ay Trade (proj)','0').replace('+Outros Custos Impostos','0').replace('Total Conta Investimento','0').replace('Venda Opções','0').replace('Taxa operacional','0').replace('Ajuste de posição','0').replace('Total Conta Normal','0').replace('Compra Opções','0').replace('Taxa registro BM&F','0').replace('Ajuste day trade','0').replace('Total liquido (#)','0').replace('Valor dos negócios','0').replace('Taxas BM&F (emol+f.gar)','0').replace('Total das despesas','0').replace('Total líquido da nota','0')
    return value

# ===================================================================================================
# Sanitinizar observação
# ===================================================================================================
def sanitiza_observacao(observacao):
    observacao = observacao.str.replace('A','').str.replace('T','').str.replace('#2','').str.replace('2','').str.replace('C','').str.replace('I','').str.replace('#','').str.replace('P','').str.replace('8','').str.replace('H','').str.replace('X','').str.replace('F','').str.replace('Y','').str.replace('B','').str.replace('L','')
    return observacao

# ===================================================================================================
# Verifica se a Nota de Corretagem já foi processada anteriormente
# ===================================================================================================
def verifica_nota_corretagem(folder_path,nome,item):
    if exists(folder_path+'/'+nome):
        print_atencao()
        print("Todas as Notas de Corretagem contidadas no arquivo",basename(item),"já foram contabilizadas anteriormente.")
        print("Caso deseje reporcessá-las se assegure de ter apagado/deletado o arquivo",nome,"na pasta",folder_path,'.')
        print("Além disso, os dados relativos a essa Nota de Corretagem e inseridos na aba DADOS do arquivo COIR.xslm também devem ser apagados.\n")
        log = 'ATENÇÃO:\n'
        log += ' - Todas as Notas de Corretagem contidadas no arquivo "'+basename(item)+'" já foram contabilizadas anteriormente.\n'
        log += ' - Caso deseje reporcessá-las se assegure de ter apagado/deletado o arquivo "'+nome+'" na pasta "'+folder_path+'".\n'
        log += ' - Além disso, os dados relativos a essa Nota de Corretagem e inseridos na aba "DADOS" do arquivo "COIR.xslm" também devem ser apagados.\n'
        return log

# ===================================================================================================
# Get Ticket - Condição 01: HASH11 - Condição 02: BABA34 - Condição 03: ITSA4 - Condição 04: USIM5
# ===================================================================================================
def get_ticket(stock_title):
    if (len(stock_title) > 5):
        if (stock_title[4] == '1') and (stock_title[5] == '1'):
            stock_title = stock_title[0:6]
        elif (stock_title[4] == '3') or (stock_title[5] == '4'):
            if (len(stock_title) > 5) and (stock_title[5] == '4'):
                stock_title = stock_title[0:6]
            else:
                stock_title = stock_title[0:5]
        elif (stock_title[4] == '4'):
            stock_title = stock_title[0:5]
        elif (stock_title[4] == '5'):
            stock_title = stock_title[0:5]
    else:
        stock_title = stock_title[0:5]
    return stock_title

# ===================================================================================================
# Altera o nome do papel no pregão para o código no padrão XXXX4, por exemplo "PETR4", "VALE3", etc.
# ===================================================================================================
def nome_pregao(acoes, stock_title, data):
    control = 0
    log = ''
    for current_row in acoes.index:
        cell_value = acoes['TICKET'].iloc[current_row]
        cell_ticket = acoes['CODIGO'].iloc[current_row]

        if cell_value == stock_title:
            stock_title = cell_ticket
            control = 1
            break
    if control == 0: # Não precisa cadastrar uma ação
        cell_ticket = stock_title[0:5]
        control = 1

    stock_title,log_altera_ticket = altera_ticket(stock_title, data)
    log += log_altera_ticket
    
    #Mensagem de alerta da variável stock_title, caso não tenha sido encontrada, conforem laço anterior
    if control == 0:
        print_atencao()
        print('O ativo',stock_title,'ainda não foi cadastrado!"')
        log += 'ATENÇÃO:\n'
        log += ' - O ativo '+stock_title+' ainda não foi cadastrado!\n'
        stock_title = stock_title[0:5]
        print('Na planilha Normais_Dados ou DayTrade_Dados do arquivo COIR.xlsb irá aparecer o código',stock_title)
        print('Realize a alteração para o código correto do ativo APENAS nessas planilhas.','\n')
        log += ' - Na planilha Normais_Dados ou DayTrade_Dados do arquivo COIR.xlsb irá aparecer o código: "'+stock_title+'".\n'
        log += ' - Realize a alteração para o código correto do ativo APENAS nessas planilhas.\n'
    return stock_title,log

# ===================================================================================================
# Obtem o valor individual das taxas de Registro e Emolumento de cada operação do mercado futuro
# Fonte: https://www.b3.com.br/pt_br/produtos-e-servicos/tarifas/consulta/
# ===================================================================================================
def taxas_registro_emol(operacao,mercadoria,stock_title):
    registro_emol = 0
    if operacao == "Normal":
        if mercadoria in "CCMccm":
            mercado = "MILHO"
            registro_emol = float(0.47 + 0.25)     
        elif mercadoria in "WINwin":
            mercado = "INDICE"
            registro_emol = float(0.25 + 0.14)   
        elif mercadoria in "WDOwdo":
            mercado = "DOLAR"
            registro_emol = float(0.73 + 0.40)
        elif mercadoria in "INDind":
            mercado = "INDICE"
            registro_emol = float(1.28 + 0.69)
        elif mercadoria in "DOLdol":
            mercado = "DOLAR"
            registro_emol = float(3.67 + 1.97)
        elif mercadoria in "BGIbgi":
            mercado = "BOI"
            registro_emol = float(1.78 + 0.96)  
        elif mercadoria in "IFCifc":
            mercado = "CAFE"
            registro_emol = float(2.63 + 1.42)
        elif mercadoria in "SJCsjc":
            mercado = "SOJA"
            registro_emol = float(2.40 + 1.39)
    elif operacao == "DayTrade":
        if mercadoria in "CCMccm":
            mercado = "MILHO"
            registro_emol = float(0.23 + 0.13)   
        elif mercadoria in "WINwin":
            mercado = "INDICE"
            registro_emol = 0.25
            #registro_emol = 0.16 + 0.09 - python 3 não aceita número iniciando com zero(0)
        elif mercadoria in "WDOwdo":
            mercado = "DOLAR"
            registro_emol = float(0.68 + 0.37)
        elif mercadoria in "INDind":
            mercado = "INDICE"
            registro_emol = float(0.83 + 0.45)
        elif mercadoria in "DOLdol":
            mercado = "DOLAR"
            registro_emol = float(3.48 + 1.88)
        elif mercadoria in "BGIbgi":
            mercado = "BOI"
            registro_emol = float(0.53 + 0.29)
        elif mercadoria in "IFCifc":
            mercado = "CAFE"
            registro_emol = float(0.79 + 0.43)
        elif mercadoria in "SJCsjc":
            mercado = "SOJA"
            registro_emol = float(2.40 + 1.39)
    else:
        mercado = stock_title
        registro_emol = 0
    return registro_emol,mercado

# ===================================================================================================
# Obtem o valor individual das taxas de Registro e Emolumento de cada operação do mercado futuro
# Dados anterior a 06/06/2022. Com atualização mais antiga em 23/01/2019
# ===================================================================================================
def taxas_registro_emol_old(operacao,mercadoria,stock_title):
    registro_emol = 0
    if operacao == "Normal":
        if mercadoria in "CCMccm":
            mercado = "MILHO"
            registro_emol = float(0.45 + 0.27)     
        elif mercadoria in "WINwin":
            mercado = "INDICE"
            registro_emol = float(0.26 + 0.13)   
        elif mercadoria in "WDOwdo":
            mercado = "DOLAR"
            registro_emol = float(0.75 + 0.41)
        elif mercadoria in "INDind":
            mercado = "INDICE"
            registro_emol = float(1.34 + 0.63)
        elif mercadoria in "DOLdol":
            mercado = "DOLAR"
            registro_emol = float(3.56 + 1.85)
        elif mercadoria in "BGIbgi":
            mercado = "BOI"
            registro_emol = float(1.47 + 1.27)  
        elif mercadoria in "IFCifc":
            mercado = "CAFE"
            registro_emol = float(2.05 + 1.71)
        elif mercadoria in "SJCsjc":
            mercado = "SOJA"
            registro_emol = float(2.41 + 1.50)
    elif operacao == "DayTrade":
        if mercadoria in "CCMccm":
            mercado = "MILHO"
            registro_emol = float(0.29 + 0.14)   
        elif mercadoria in "WINwin":
            mercado = "INDICE"
            registro_emol = 0.25
            #registro_emol = 0.06 + 0.07 - python 3 não aceita número iniciando com zero(0)
        elif mercadoria in "WDOwdo":
            mercado = "DOLAR"
            registro_emol = float(0.71 + 0.39)
        elif mercadoria in "INDind":
            mercado = "INDICE"
            registro_emol = float(0.91 + 0.41)
        elif mercadoria in "DOLdol":
            mercado = "DOLAR"
            registro_emol = float(3.39 + 1.76)
        elif mercadoria in "BGIbgi":
            mercado = "BOI"
            registro_emol = float(0.53 + 0.38)
        elif mercadoria in "IFCifc":
            mercado = "CAFE"
            registro_emol = float(0.63 + 0.62)
        elif mercadoria in "SJCsjc":
            mercado = "SOJA"
            registro_emol = float(2.41 + 1.50)
    else:
        mercado = stock_title
        registro_emol = 0
    return registro_emol,mercado

# ===================================================================================================
# Altera o nome do papel no pregão para o código no padrão XXXX4, por exemplo "PETR4", "VALE3", etc.
# ===================================================================================================
def nome_pregao_opcoes(opcoes, stock_title, data):
    control = 0
    log = ''
    for current_row in opcoes.index:
        cell_value = opcoes['TICKET'].iloc[current_row]
        if cell_value in stock_title:
            stock_title = opcoes['CODIGO'].iloc[current_row]
            control = 1
            break
    stock_title,log_altera_ticket = altera_ticket(stock_title, data)
    log += log_altera_ticket
    
    #Mensagem de alerta da variável stock_title, caso não tenha sido encontrada, conforem laço anterior
    if control == 0:
        print_atencao()
        #print('\n')
        print('O ativo',stock_title,'ainda não foi cadastrado!"')
        log += 'ATENÇÃO:\n'
        log += ' - O ativo '+stock_title+' ainda não foi cadastrado!\n'
        stock_title = stock_title[0:5] + '99'
        print('Na planilha Normais_Dados ou DayTrade_Dados do arquivo COIR.xlsb irá aparecer o código',stock_title)
        print('Realize a alteração para o código correto do ativo APENAS nessas planilhas.','\n')
        log += ' - Na planilha Normais_Dados ou DayTrade_Dados do arquivo COIR.xlsb irá aparecer o código: "'+stock_title+'".\n'
        log += ' - Realize a alteração para o código correto do ativo APENAS nessas planilhas.\n'
    return stock_title,log    
    
# ===================================================================================================
# Altera o nome de papel conforme as atualizações das empresas na B3
# ===================================================================================================
def altera_ticket(stock, data):
    log = ''
    if stock == "JSLG3":
        datalimite = datetime.strptime("19/09/2020", '%d/%m/%Y').date()
        if data < datalimite:
            stock = "SIMH3"
            log = 'As operações anteriores ao dia 19/09/2020 com o código "JSLG3" foram alteradas para o código "SIMH3".\n'
    return stock,log

# ===================================================================================================
# Obtem o valor total de cada operação, o multipicador e o código B3.
# ===================================================================================================    
def mercadoria_ticket(mercadoria, preco_unitario, quantidade):        
    if 'CCM' in mercadoria:
        valor_total = preco_unitario * quantidade * 450
        id = 'CCM'
        mult = 450
    elif 'BGI' in mercadoria:
        valor_total = preco_unitario * quantidade * 330
        id = 'BGI'
        mult = 330
    elif 'ICF' in mercadoria: #cotação variável conforme o dolar
        valor_total = preco_unitario * quantidade * 100 # * dolar
        id = 'ICF'
        mult = 100
        print_atencao()
        print('A Commodity ',mercadoria,' ainda não teve seus dados testados exaustivamente!"')
        print('É importante verificar nas planilhas de operações se os valores estão corretos.')
        log += 'ATENÇÃO:\n'
        log += ' - A Commodity ',mercadoria,' ainda não teve seus dados testados exaustivamente!\n'
        log += ' - É importante verificar nas planilhas de operações se os valores estão corretos.\n'
    elif 'SJC' in mercadoria: #cotação variável conforme o dolar
        valor_total = preco_unitario * quantidade * 450 # * dolar
        id = 'SJC'
        mult = 450
        print_atencao()
        print('A Commodity ',mercadoria,' ainda não teve seus dados testados exaustivamente!"')
        print('É importante verificar nas planilhas de operações se os valores estão corretos.')
        log += 'ATENÇÃO:\n'
        log += ' - A Commodity ',mercadoria,' ainda não teve seus dados testados exaustivamente!\n'
        log += ' - É importante verificar nas planilhas de operações se os valores estão corretos.\n'
    elif 'WIN' in mercadoria:
        valor_total = preco_unitario * quantidade * 0.2
        id = 'WIN'
        mult = 0.2
    elif 'IND' in mercadoria:
        valor_total = preco_unitario * quantidade * 1
        id = 'IND'
        mult = 1
    elif 'WDO' in mercadoria:
        valor_total = preco_unitario * quantidade * 10
        id = 'WDO'
        mult = 10
    elif 'DOL' in mercadoria:
        valor_total = preco_unitario * quantidade * 50
        id = 'DOL'
        mult = 50
    return valor_total,id,mult

# ===================================================================================================
#Contabiliza a quantidade de vendas e o valor de IR nas operações DayTrade e Normal BM&F
# ===================================================================================================  
def ir_bmf(cont_notas,note_df,taxas_df,row_data,note_data):
    cont_note = len(note_df['Nota'])
    
    for t in range(0,cont_notas):
        cont_v_daytrade = 0
        cont_v_normal = 0
        cont_c_normal = 0
        if taxas_df['IRRF'].iloc[t] > 0:
            for n in range(0,cont_note):
                if taxas_df['Nota'].iloc[t] == note_df['Nota'].iloc[n] and note_df['Operacao'].iloc[n] == 'Normal' and note_df['C/V'].iloc[n] == 'V':
                    cont_v_normal += 1
            if cont_v_normal == 0:
                for n in range(0,cont_note):
                    if taxas_df['Nota'].iloc[t] == note_df['Nota'].iloc[n] and note_df['Operacao'].iloc[n] == 'Normal' and note_df['C/V'].iloc[n] == 'C':
                        cont_c_normal += 1
            if cont_v_normal > 0:
                for n in range(0,cont_note):
                    if taxas_df['Nota'].iloc[t] == note_df['Nota'].iloc[n] and note_df['Operacao'].iloc[n] == 'Normal' and note_df['C/V'].iloc[n] == 'V':
                        #note_df['IRRF'].iloc[n] = taxas_df['IRRF'].iloc[t] / cont_v_normal
                        row_data = [note_df['Corretora'].iloc[n],note_df['CPF'].iloc[n],note_df['Nota'].iloc[n],note_df['Data'].iloc[n],note_df['C/V'].iloc[n],note_df['Papel'].iloc[n],note_df['Operacao'].iloc[n],note_df['Preço'].iloc[n],0,0,0,0,taxas_df['IRRF'].iloc[t]/cont_v_normal, 0,0,0,note_df['Mercado'].iloc[n]]
                        note_data.append(row_data)
            elif cont_c_normal > 0:
                for n in range(0,cont_note):
                    if taxas_df['Nota'].iloc[t] == note_df['Nota'].iloc[n] and note_df['Operacao'].iloc[n] == 'Normal' and note_df['C/V'].iloc[n] == 'C':
                        #note_df['IRRF'].iloc[n] = taxas_df['IRRF'].iloc[t] / cont_c_normal
                        row_data = [note_df['Corretora'].iloc[n],note_df['CPF'].iloc[n],note_df['Nota'].iloc[n],note_df['Data'].iloc[n],note_df['C/V'].iloc[n],note_df['Papel'].iloc[n],note_df['Operacao'].iloc[n],note_df['Preço'].iloc[n],0,0,0,0,taxas_df['IRRF'].iloc[t]/cont_c_normal, 0,0,0,note_df['Mercado'].iloc[n]]
                        note_data.append(row_data)
        elif taxas_df['IR_DT'].iloc[t] > 0:
            for n in range(0,cont_note):
                if taxas_df['Nota'].iloc[t] == note_df['Nota'].iloc[n] and note_df['Operacao'].iloc[n] == 'DayTrade' and note_df['C/V'].iloc[n] == 'V':
                    cont_v_daytrade += 1
            for n in range(0,cont_note):
                if taxas_df['Nota'].iloc[t] == note_df['Nota'].iloc[n] and note_df['Operacao'].iloc[n] == 'DayTrade' and note_df['C/V'].iloc[n] == 'V':
                    #note_df['IRRF'].iloc[n] = taxas_df['IR_DT'].iloc[t] / cont_v_daytrade
                    '''Testar com essa nova substituição de campo'''
                    #note_df['IR_DT'].iloc[n] = taxas_df['IR_DT'].iloc[t] / cont_v_daytrade
                    row_data = [note_df['Corretora'].iloc[n],note_df['CPF'].iloc[n],note_df['Nota'].iloc[n],note_df['Data'].iloc[n],note_df['C/V'].iloc[n],note_df['Papel'].iloc[n],note_df['Operacao'].iloc[n],note_df['Preço'].iloc[n],0,0,0,0,taxas_df['IR_DT'].iloc[t]/cont_v_daytrade, 0,0,0,note_df['Mercado'].iloc[n]]
                    note_data.append(row_data)
    cols = ['Corretora','CPF', 'Nota', 'Data', 'C/V', 'Papel', 'Operacao','Preço','Quantidade', 'Total','Custos_Fin','PM','IRRF','IR_DT','ID','FATOR','Mercado']
    note_df = pd.DataFrame(data=note_data, columns=cols)  
    #print(note_df.head(40))
    return(note_df)

# ===================================================================================================
# Cria o caminho completo de pasta/subpasta para salvar o resultado do processamento
# ===================================================================================================
def move_resultado(folder_path,cpf,nome,item,pagebmf):
    from os import makedirs
    from shutil import copytree, ignore_patterns
    if exists(folder_path+'/'+nome):
        print_atencao()
        print("Todas as Notas de Corretagem contidadas no arquivo",basename(item),"já foram contabilizadas anteriormente.")
        print("Caso deseje reporcessá-las se assegure de ter apagado/deletado o arquivo",nome,"na pasta",folder_path,'.')
        print("Além disso, os dados relativos a essa Nota de Corretagem e inseridos na aba DADOS do arquivo COIR.xslm também devem ser apagados.\n")
        log = 'ATENÇÃO:\n'
        log += ' - Todas as Notas de Corretagem contidadas no arquivo "'+basename(item)+'" já foram contabilizadas anteriormente.\n'
        log += ' - Caso deseje reporcessá-las se assegure de ter apagado/deletado o arquivo "'+nome+'" na pasta "'+folder_path+'".\n'
        log += ' - Além disso, os dados relativos a essa Nota de Corretagem e inseridos na aba "DADOS" do arquivo "COIR.xslm" também devem ser apagados.\n'
        pagebmf = 0
    elif not exists('./Resultado/'+cpf):
        source = './Apoio/'
        destination = './Resultado/'+cpf
        copytree(source, destination, ignore=ignore_patterns('*.py', '*.pyc', 'tmp*','*.csv','Backup','_Temp','Screenshots'))
        makedirs(folder_path)
        log = ''
    elif not exists(folder_path):
        makedirs(folder_path)
        log = ''
    else:
        log = ''
    return log,pagebmf

# ===================================================================================================
# Agrupar os dados de preço e quantidade por cada ativo comprado/vendido em cada nota de corretagem
# ===================================================================================================
def agrupar_bmf(note_df):
    note_df_agrupado = note_df.groupby(
    ['Corretora','CPF','Nota', 'Data', 'C/V', 'Papel', 'Operacao','Preço','Mercado'],as_index=False
    ).agg(
        {
            'Quantidade': sum,
            'Total': sum,
            'Custos_Fin': sum,
            'PM': sum,
            'IRRF': sum,
            'IR_DT': sum
        }
    )
    return note_df_agrupado

# ===================================================================================================
# Agrupar os dados de preço e quantidade por cada ativo comprado/vendido em cada nota de corretagem
# ===================================================================================================
def agrupar(note_df):
    note_df_agrupado = note_df.groupby(
    ['Corretora','CPF','Nota', 'Data', 'C/V', 'Papel', 'Operacao','Mercado','Prazo','Exercicio'],as_index=False
    ).agg(
        {
            'Preço': sum,
            'Quantidade': sum,
            'Total': sum,
            'Custos_Fin': sum,
            'PM': sum,
            'IRRF': sum
        }
    )
    return note_df_agrupado

# ===================================================================================================
# Acrescentando os custos operacionais (Corretagem, Imposto e Outros)
# ===================================================================================================
def custos_operacionais(grouped,taxas_df):
    #grouped['freq'] = grouped.groupby('Nota')['Nota'].transform('count')
    custos_data = []
    for current_row in taxas_df.index:
        for I in grouped.index:
            if grouped['Nota'].iloc[I] == taxas_df['Nota'].iloc[current_row]:
                row_data = float(taxas_df['Custos_Op'].iloc[current_row]) * (float(grouped['Total'].iloc[I])/float(taxas_df['Total'].iloc[current_row]))
                custos_data.append(row_data)
    grouped['Custos_Op'] = pd.DataFrame(data=custos_data)
    grouped['Custos_Fin'] = grouped['Custos_Fin'] + grouped['Custos_Op']
    return grouped

# ===================================================================================================
# Acrescentando os custos operacionais (Corretagem, Imposto e Outros) Operações BMF
# ===================================================================================================
def custos_financeiros(grouped,taxas_df):
    grouped['freq'] = grouped.groupby('Nota')['Nota'].transform('count')
    custos_data = []
    for current_row in taxas_df.index:
        for I in grouped.index:
            if grouped['Nota'].iloc[I] == taxas_df['Nota'].iloc[current_row]:
                row_data = float(taxas_df['Custos_Fin'].iloc[current_row])/float(grouped['freq'].iloc[I])
                custos_data.append(row_data)
    grouped['Custos_Fin'] = pd.DataFrame(data=custos_data)
    return grouped

# ===================================================================================================
# Agrupa as operações por tipo de trade (normal ou daytrade) com correção 
# de compra/venda a maior no DayTrade
# ===================================================================================================
def agrupar_operacoes_correcao(grouped,cols):
    groups = grouped.groupby(grouped.Operacao)
    try:
        normal_df = groups.get_group("Normal")
    except KeyError:
        normal_df = pd.DataFrame(columns=cols)
    try:
        daytrade_df = groups.get_group("DayTrade")
    except KeyError:
        daytrade_df = pd.DataFrame(columns=cols)
    return normal_df,daytrade_df

# ===================================================================================================
# Obtendo o valor correto do preço médio de cada operação
# ===================================================================================================
def preco_medio_correcao(grouped):
    preco_medio_data = []
    for current_row in grouped.index:
        if grouped['C/V'].iloc[current_row] == 'C':
            preco_medio = (grouped['Total'].iloc[current_row] + grouped['Custos_Fin'].iloc[current_row]) / grouped['Quantidade'].iloc[current_row]
        else:
            preco_medio = (grouped['Total'].iloc[current_row] - grouped['Custos_Fin'].iloc[current_row]) / grouped['Quantidade'].iloc[current_row]
        row_data = [preco_medio]
        preco_medio_data.append(row_data)
    cols = ['PM']
    preco_medio_df = pd.DataFrame(data=preco_medio_data, columns=cols)
    return preco_medio_df['PM']

# ===================================================================================================
# Formata valores floats para uma quantidade específica de casas decimais
# ===================================================================================================
def truncate(number, decimals=0):
    from math import trunc
    if decimals < 0:
        raise ValueError('truncate received an invalid value of decimals ({})'.format(decimals))
    elif decimals == 0:
        return trunc(number)
    else:
        factor = float(10**decimals)
        return trunc(number*factor)/factor

'''Os valores de PM e total dos sistemas comerciais não têm limite de casas decimais.
Após teste de controle foi observado que TRUNCAR o PM na planilha COIR ficava com alguns centavos 
a menor ou maior em relação a esses sistemas'''
        
# ===================================================================================================
# Quantidade operada de cada ativo por nota de corretagem
# ===================================================================================================
def quantidade_operada(df_quantidade=0,df_unnamed_0=0,df_unnamed_1=0,df_unnamed_2=0):
    quantidade = 0
    if df_quantidade > 0:
        quantidade = df_quantidade
    elif isinstance(df_unnamed_0, float) and df_unnamed_0 > 0:
        quantidade = df_unnamed_0
    elif df_unnamed_1 > 0:
        quantidade = df_unnamed_1
    elif df_unnamed_2 > 0:
        quantidade = df_unnamed_2
    return quantidade

# ===================================================================================================
# Valor total de cada operação por nota de corretagem
# ===================================================================================================
def valor_total_ativo(df_valor_operacao_ajuste,df_unnamed_2=0,df_unnamed_1=0):
    total = 0
    if df_valor_operacao_ajuste > 0:
        total = df_valor_operacao_ajuste
    elif df_unnamed_2 > 0:
        total = df_unnamed_2
    elif df_unnamed_1 > 0:
        total = df_unnamed_1
    return total

# ===================================================================================================
# Obter o preço unitário de cada operação
# ===================================================================================================
'''Foi excluída a etapa para extração do preço unitário de cada operação.
   Pois esse dado é obtido pela divisão do valor_total / quantidade.'''
#def obter_preco():
#    if df['Preço / Ajuste'].iloc[current_row] > 0:
#        price = df['Preço / Ajuste'].iloc[current_row]
#    elif 'Unnamed: 0' in df.columns:
#        if df['Unnamed: 0'].iloc[current_row] > 0 and df['Quantidade'].iloc[current_row] > 0:
#            price = df['Unnamed: 0'].iloc[current_row]
#        elif 'Unnamed: 1' in df.columns:
#            if df['Unnamed: 1'].iloc[current_row] > 0:
#                price = df['Unnamed: 1'].iloc[current_row]
#            elif 'Unnamed: 2' in df.columns:
#                if df['Unnamed: 2'].iloc[current_row] > 0:
#                    price = df['Unnamed: 2'].iloc[current_row]

# ===================================================================================================
# Dividindo os custos operacionais e o IRRF por cada operação.
# ===================================================================================================
def custos_por_operacao(taxas_df,number,c_v,total,operacao):
    custos_fin = 0
    irrf = 0
    for current_row in taxas_df.index:
        cell_value = taxas_df['Nota'].iloc[current_row]
        if cell_value == number:
            custos_fin = (total/taxas_df['Total'].iloc[current_row])*taxas_df['Custos_Fin'].iloc[current_row]
            if c_v == 'V' and operacao == "Normal":
                if taxas_df['BaseCalculo'].iloc[current_row] > 0:
                    irrf = (total/taxas_df['BaseCalculo'].iloc[current_row])*taxas_df['IRRF'].iloc[current_row]
                else:
                    irrf = (total/taxas_df['Vendas'].iloc[current_row])*taxas_df['IRRF'].iloc[current_row]
            else:
                irrf = 0
    return custos_fin,irrf

# ===================================================================================================
# Dividindo os custos operacionais e o IRRF por cada operação BMF e Contratos Futuros.
# ================================================================================================
def custos_por_operacao_bmf(taxas_df,number,corretagem):
    custos_fin = 0
    for current_row in taxas_df.index:
        cell_value = taxas_df['Nota'].iloc[current_row]
        if cell_value == number:
            custos_fin = (corretagem/taxas_df['Corretagem'].iloc[current_row])*taxas_df['Custos_Fin'].iloc[current_row]
    return custos_fin
    
# ===================================================================================================
# Converte o nome de uma opção de Comra/Venda em em um código no padrão XXXX4
# cods  = {'ON' : '3' , 'PN' : '4' , 'PNA' : '5' , 'PNB' : '6' , 'PNC' : '7' , 'UNT' : '11'}
# ===================================================================================================
def converte_opcao_ticket(opcao):
    ativo = opcao.split(" ")[0][0:4]
    try:
        letras = opcao.split(" ")[1]
        codigo  =  Apoio.tickets.codigos[letras]
    except:
        codigo = '11'
    ativo = ativo + codigo
    return ativo

# ===================================================================================================
# Obtendo o valor preço médio de cada operação
# ===================================================================================================
def preco_medio(c_v,total,custos_fin,quantidade):
    preco_medio = 0
    if c_v == "C":
        preco_medio = round((total + custos_fin)/quantidade,4)
    elif c_v =="V":
        preco_medio = round((total - custos_fin)/quantidade,4)
    return preco_medio
    
# ===================================================================================================
# Validação de notas de corretagem no padrão Sinacor
# ===================================================================================================
def valida_nota_corretagem(validacao,item): 
    df_validacao = pd.concat(validacao,axis=1,ignore_index=True)
    df_validacao = pd.DataFrame({'NotaCorretagem': df_validacao[0].unique()})
    cell_value = df_validacao['NotaCorretagem'].iloc[0]
    if cell_value == 'NOTA DE NEGOCIAÇÃO' or 'NOTA DE CORRETAGEM':
        print('processando o arquivo:',basename(item))
    #return()
    '''Função não está em uso
    os comando de continue não pode ser utilizados fora de laços'''

# ===================================================================================================
# Validação de notas de corretagem no padrão Sinacor
# ===================================================================================================
def valida_corretora(corretora):
    control = 0
    df_corretora = pd.concat(corretora,axis=1,ignore_index=True)
    cell_value = df_corretora[1].iloc[3]
    cell_value = cell_value.upper()

    if type(df_corretora[0].iloc[3]) == str:
        nota_BMF = df_corretora[0].iloc[3]
        nota_BMF = nota_BMF.upper()
    for current_row in Apoio.tickets.corretoras_cadastradas.index:
        corretora_value = Apoio.tickets.corretoras_cadastradas['Corretora'].iloc[current_row]
        if cell_value == corretora_value:
            corretora = Apoio.tickets.corretoras_cadastradas['Nome'].iloc[current_row]
            control = 1
            break    
    if control == 0:
        for current_row in Apoio.tickets.corretoras_cadastradas.index:
            corretora_value = Apoio.tickets.corretoras_cadastradas['Corretora'].iloc[current_row]
            if nota_BMF == corretora_value:
                corretora = Apoio.tickets.corretoras_cadastradas['Nome'].iloc[current_row]
                control = 2
                break
    return control,corretora,cell_value

# ===================================================================================================
# Agrupar os dados de preço e quantidade por cada ativo comprado/vendido em cada nota de corretagem
# ===================================================================================================
def agrupar_old(note_df):
    note_df_agrupado = note_df.groupby(
    ['Corretora','CPF','Nota', 'Data', 'C/V', 'Papel', 'Operacao'],as_index=False
    ).agg(
        {
            'Preço': sum,
            'Quantidade': sum,
            'Total': sum,
            'Custos_Fin': sum,
            'PM': sum,
            'IRRF': sum
        }
    )
    return note_df_agrupado

# ===================================================================================================
# Seleção de papel isento de IR (IRRF e IRPF)
# Caso haja mais de um papel isento em uma mesma NC o sistema NÃO detectará
# A implementação de analisar mais de um papel isento por NC continua pendente
# ===================================================================================================     
def isencao_imposto_renda(taxas_df,grouped,note_data):
    log = ''
    controle = 0
    #Analisa as operações com uma venda isenta de IR
    for current_row in taxas_df.index:
        if (taxas_df['Vendas'].iloc[current_row] and taxas_df['BaseCalculo'].iloc[current_row]) > 0 and taxas_df['Vendas'].iloc[current_row] != taxas_df['BaseCalculo'].iloc[current_row]:
            diff = taxas_df['Vendas'].iloc[current_row] - taxas_df['BaseCalculo'].iloc[current_row]
            for I in range(0,len(grouped)):
                if grouped['Total'].iloc[I] == diff and grouped['Operacao'].iloc[I] == 'Normal':
                    note_data.append([grouped['Corretora'].iloc[I],grouped['CPF'].iloc[I],grouped['Nota'].iloc[I],grouped['Data'].iloc[I],'V',grouped['Papel'].iloc[I],'Normal',0,0,0,0,0,(grouped['IRRF'].iloc[I])*-1,grouped['Mercado'].iloc[I],grouped['Prazo'].iloc[I],grouped['Exercicio'].iloc[I]])
                    datetime.today().strftime('%d/%m/%Y')
                    log = 'Na operação de venda de "'+str(grouped["Papel"].iloc[I])+'", nota de corretagem nº '+str(int(grouped["Nota"].iloc[I]))+', do dia '+str(grouped["Data"].iloc[I].strftime('%d/%m/%Y'))+', houve isenção de Imposto de Renda.\n'
                    controle = 1   
            #Analisa as operações com duas vendas isentas de IR
            if controle == 0:
                for i in range(0,len(grouped)):
                    for p in range(grouped.index[i],len(grouped)):
                        if grouped['Total'].iloc[i] + grouped['Total'].iloc[p] == diff and grouped['Operacao'].iloc[i] == 'Normal':
                            note_data.append([grouped['Corretora'].iloc[i],grouped['CPF'].iloc[i],grouped['Nota'].iloc[i],grouped['Data'].iloc[i],'V',grouped['Papel'].iloc[i],'Normal',0,0,0,0,0,(grouped['IRRF'].iloc[i])*-1,grouped['Mercado'].iloc[i],grouped['Prazo'].iloc[i],grouped['Exercicio'].iloc[i]])
                            note_data.append([grouped['Corretora'].iloc[p],grouped['CPF'].iloc[p],grouped['Nota'].iloc[p],grouped['Data'].iloc[p],'V',grouped['Papel'].iloc[p],'Normal',0,0,0,0,0,(grouped['IRRF'].iloc[p])*-1,grouped['Mercado'].iloc[p],grouped['Prazo'].iloc[p],grouped['Exercicio'].iloc[p]])
                            datetime.today().strftime('%d/%m/%Y')
                            log = 'Na operação de venda de "'+str(grouped["Papel"].iloc[i])+'", nota de corretagem nº '+str(int(grouped["Nota"].iloc[i]))+', do dia '+str(grouped["Data"].iloc[i].strftime('%d/%m/%Y'))+', houve isenção de Imposto de Renda.\n'
                            log += 'Na operação de venda de "'+str(grouped["Papel"].iloc[p])+'", nota de corretagem nº '+str(int(grouped["Nota"].iloc[p]))+', do dia '+str(grouped["Data"].iloc[p].strftime('%d/%m/%Y'))+', houve isenção de Imposto de Renda.\n'
                            controle = 1
                            break
                    else:
                        continue
                    break
            #Analisa as operações com três vendas isentas de IR
            if controle == 0:
                for i in range(0,len(grouped)):
                    for p in range(grouped.index[i],len(grouped)):
                        for z in range(grouped.index[p],len(grouped)):
                            if grouped['Total'].iloc[i] + grouped['Total'].iloc[p] + grouped['Total'].iloc[z] == diff and grouped['Operacao'].iloc[i] == 'Normal':
                                note_data.append([grouped['Corretora'].iloc[i],grouped['CPF'].iloc[i],grouped['Nota'].iloc[i],grouped['Data'].iloc[i],'V',grouped['Papel'].iloc[i],'Normal',0,0,0,0,0,(grouped['IRRF'].iloc[i])*-1,grouped['Mercado'].iloc[i],grouped['Prazo'].iloc[i],grouped['Exercicio'].iloc[i]])
                                note_data.append([grouped['Corretora'].iloc[p],grouped['CPF'].iloc[p],grouped['Nota'].iloc[p],grouped['Data'].iloc[p],'V',grouped['Papel'].iloc[p],'Normal',0,0,0,0,0,(grouped['IRRF'].iloc[p])*-1,grouped['Mercado'].iloc[p],grouped['Prazo'].iloc[p],grouped['Exercicio'].iloc[p]])
                                note_data.append([grouped['Corretora'].iloc[z],grouped['CPF'].iloc[z],grouped['Nota'].iloc[z],grouped['Data'].iloc[z],'V',grouped['Papel'].iloc[z],'Normal',0,0,0,0,0,(grouped['IRRF'].iloc[z])*-1,grouped['Mercado'].iloc[z],grouped['Prazo'].iloc[z],grouped['Exercicio'].iloc[z]])
                                datetime.today().strftime('%d/%m/%Y')
                                log = 'Na operação de venda de "'+str(grouped["Papel"].iloc[i])+'", nota de corretagem nº '+str(int(grouped["Nota"].iloc[i]))+', do dia '+str(grouped["Data"].iloc[i].strftime('%d/%m/%Y'))+', houve isenção de Imposto de Renda.\n'
                                log += 'Na operação de venda de "'+str(grouped["Papel"].iloc[p])+'", nota de corretagem nº '+str(int(grouped["Nota"].iloc[p]))+', do dia '+str(grouped["Data"].iloc[p].strftime('%d/%m/%Y'))+', houve isenção de Imposto de Renda.\n'
                                log += 'Na operação de venda de "'+str(grouped["Papel"].iloc[z])+'", nota de corretagem nº '+str(int(grouped["Nota"].iloc[z]))+', do dia '+str(grouped["Data"].iloc[z].strftime('%d/%m/%Y'))+', houve isenção de Imposto de Renda.\n'
                                controle = 1                               
                                break
                        else:
                            continue
                        break
                    else:
                        continue
                    break
    return note_data,log

# ===================================================================================================
# Agrupa as operações por tipo de operação (Normal ou Daytrade)
# ===================================================================================================
def agrupar_operacoes(grouped,cols):
    groups = grouped.groupby(grouped.Operacao)
    try:
        normal_df = groups.get_group("Normal")
    except KeyError:
        normal_df = pd.DataFrame(columns=cols)
    try:
        daytrade_df = groups.get_group("DayTrade")
        vendas = daytrade_df.loc[daytrade_df['C/V'] == 'V']
        compras = daytrade_df.loc[daytrade_df['C/V'] == 'C']
        result = pd.merge(compras, vendas, on=["Corretora", "CPF","Nota","Data","Papel",'Operacao','IRRF'])
        result['QTDE'] = result['Quantidade_x'] - result['Quantidade_y']
        result['Lucro'] = (result['Total_y']/result['Quantidade_y'] - result['Total_x']/result['Quantidade_x'])*((result['QTDE']+result['Quantidade_x']+result['Quantidade_y'])/2)
        return normal_df,daytrade_df,result
    except KeyError:
        daytrade_df = pd.DataFrame(columns=cols)
        return normal_df,daytrade_df

# ===================================================================================================
# Insere o valor do IR para as operações de Daytrade"
# ===================================================================================================
def daytrade_ir(result,taxas_df,note_data,grouped):
    log = ''
    result = result
    for current_row in result.index:
        if result['Lucro'].iloc[current_row] > 0:
            for I in taxas_df.index:
                if taxas_df['Nota'].iloc[I] == result['Nota'].iloc[current_row] and taxas_df['IR_DT'].iloc[I] > 0:
                    row_data = [result['Corretora'].iloc[current_row],result['CPF'].iloc[current_row],result['Nota'].iloc[current_row],result['Data'].iloc[current_row],'V',result['Papel'].iloc[current_row],'DayTrade',0,0,0,0,0,taxas_df['IR_DT'].iloc[I]]
                    note_data.append(row_data)
                    break
    # ===================================================================================================
    # As operações de DayTrade com sobras deverão ter esse excesso adicionados nas operações Normais
    # Por exemplo: Compra 600 - Venda 400. 200 será inserido em uma operação Normal 
    # ===================================================================================================           
    for i in result.index:
        qtde = result['QTDE'].iloc[i]
        nota = result['Nota'].iloc[i]
        papel = result['Papel'].iloc[i]
        if qtde > 0:
            for I in grouped.index:
                if grouped['Nota'].iloc[I] == nota and grouped['Papel'].iloc[I] == papel and grouped['Operacao'].iloc[I] == 'DayTrade' and grouped['C/V'].iloc[I] == 'C':
                    row_data = [grouped['Corretora'].iloc[I],grouped['CPF'].iloc[I],grouped['Nota'].iloc[I],grouped['Data'].iloc[I],'C',grouped['Papel'].iloc[I],'DayTrade',0,qtde*(-1),(grouped['Total'].iloc[I]/grouped['Quantidade'].iloc[I])*qtde*(-1),(grouped['Custos_Fin'].iloc[I]/grouped['Quantidade'].iloc[I])*qtde*(-1),0,0,grouped['Mercado'].iloc[I],grouped['Prazo'].iloc[I],grouped['Exercicio'].iloc[I]]
                    note_data.append(row_data)
                    row_data = [grouped['Corretora'].iloc[I],grouped['CPF'].iloc[I],grouped['Nota'].iloc[I],grouped['Data'].iloc[I],'C',grouped['Papel'].iloc[I],'Normal',grouped['Total'].iloc[I]/grouped['Quantidade'].iloc[I],qtde,(grouped['Total'].iloc[I]/grouped['Quantidade'].iloc[I])*qtde,(grouped['Custos_Fin'].iloc[I]/grouped['Quantidade'].iloc[I])*qtde,0,0,grouped['Mercado'].iloc[I],grouped['Prazo'].iloc[I],grouped['Exercicio'].iloc[I]]
                    note_data.append(row_data)
                    print(f'Houve operação de Daytrade do papel {papel} com sobra de {qtde} açoes para Swing Trade, na operação de COMPRA.')
                    log += 'Houve operação de Daytrade do papel "'+papel+'" com sobra de "'+str(int(qtde))+'" açoes para Swing Trade, na operação de COMPRA.\n'
                    break
        elif qtde < 0:
            qtde = qtde*(-1)
            for I in grouped.index:
                if grouped['Nota'].iloc[I] == nota and grouped['Papel'].iloc[I] == papel and grouped['Operacao'].iloc[I] == 'DayTrade' and grouped['C/V'].iloc[I] == 'V':
                    row_data = [grouped['Corretora'].iloc[I],grouped['CPF'].iloc[I],grouped['Nota'].iloc[I],grouped['Data'].iloc[I],'V',grouped['Papel'].iloc[I],'DayTrade',0,qtde*(-1),(grouped['Total'].iloc[I]/grouped['Quantidade'].iloc[I])*qtde*(-1),(grouped['Custos_Fin'].iloc[I]/grouped['Quantidade'].iloc[I])*qtde*(-1),0,0,grouped['Mercado'].iloc[I],grouped['Prazo'].iloc[I],grouped['Exercicio'].iloc[I]]
                    note_data.append(row_data)
                    row_data = [grouped['Corretora'].iloc[I],grouped['CPF'].iloc[I],grouped['Nota'].iloc[I],grouped['Data'].iloc[I],'V',grouped['Papel'].iloc[I],'Normal',grouped['Total'].iloc[I]/grouped['Quantidade'].iloc[I],qtde,(grouped['Total'].iloc[I]/grouped['Quantidade'].iloc[I])*qtde,(grouped['Custos_Fin'].iloc[I]/grouped['Quantidade'].iloc[I])*qtde,0,((grouped['Total'].iloc[I]/grouped['Quantidade'].iloc[I])*qtde)*0.00005,grouped['Mercado'].iloc[I],grouped['Prazo'].iloc[I],grouped['Exercicio'].iloc[I]]
                    note_data.append(row_data) 
                    print(f'Houve operação de Daytrade do papel {papel} com sobra de {qtde} açoes para Swing Trade, na operação de VENDA.')
                    log += 'Houve operação de Daytrade do papel "'+papel+'" com sobra de "'+str(int(qtde))+'" açoes para Swing Trade, na operação de VENDA.\n'
                    break
    return note_data,taxas_df,log

# ===================================================================================================
# Disponibiliza os dados coletados em um arquivo .xlsx separado por mês para as operações à vista
# ===================================================================================================
def arquivo_separado(folder_path,nome,note_df,normal_df,daytrade_df,taxas_df):
    with pd.ExcelWriter(folder_path+'/'+nome) as writer:
        note_df.to_excel(writer, sheet_name="Completo", index = False, merge_cells=True)
        normal_df.to_excel(writer, sheet_name="Normal", index = False, merge_cells=True)
        daytrade_df.to_excel(writer, sheet_name="DayTrade", index = False, merge_cells=True)
        taxas_df.to_excel(writer, sheet_name="Taxas", index = False, merge_cells=True)

# ===================================================================================================
# Disponibiliza os dados coletados em um arquivo .xlsx separado por mês para as operações BM&F
# ===================================================================================================
def arquivo_separado_bmf(folder_path,nome,note_df,normal_df,daytrade_df,taxas_df):
    from openpyxl import load_workbook
    book = load_workbook(folder_path+'/'+nome)
    writer = pd.ExcelWriter(folder_path+'/'+nome, engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    note_df.to_excel(writer, sheet_name="Completo", startrow=writer.sheets['Completo'].max_row, index = False,header= False)
    normal_df.to_excel(writer, sheet_name="Normal", startrow=writer.sheets['Normal'].max_row, index = False,header= False)
    daytrade_df.to_excel(writer, sheet_name="DayTrade", startrow=writer.sheets['DayTrade'].max_row, index = False,header= False)
    taxas_df.to_excel(writer, sheet_name="Taxas", startrow=writer.sheets['Taxas'].max_row, index = False,header= False)
    writer.save()

# ===================================================================================================
# Disponibiliza todos os dados coletados de todos os arquivos processados em um único arquivo
# ===================================================================================================
def arquivo_unico(current_path,cpf,note_df,normal_df,daytrade_df,taxas_df):
    from openpyxl import load_workbook
    import xlwings as xw
    log = ''
    try:
        book = load_workbook(current_path+cpf+"/Completo.xlsx")
        writer = pd.ExcelWriter(current_path+cpf+"/Completo.xlsx", engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        note_df.to_excel(writer, sheet_name="Completo", startrow=writer.sheets['Completo'].max_row, index = False,header= False)
        normal_df.to_excel(writer, sheet_name="Normal", startrow=writer.sheets['Normal'].max_row, index = False,header= False)
        daytrade_df.to_excel(writer, sheet_name="DayTrade", startrow=writer.sheets['DayTrade'].max_row, index = False,header= False)
        taxas_df.to_excel(writer, sheet_name="Taxas", startrow=writer.sheets['Taxas'].max_row, index = False,header= False)
        writer.save()
   
        # ===================================================================================================
        # Utilização do pacote xlwings para manipular arquivos excel com macro. 
        # O openpyxl se mostrou ineficaz para essa tarefa.
        # https://www.dataquest.io/blog/python-excel-xlwings-tutorial/
        # ===================================================================================================                   
        app = xw.App(visible=False)
        wb = xw.Book(current_path+cpf+"/COIR.xlsb")                 
        
        # ===================================================================================================
        # Desproteger as planilhas no arquivo COIR.xlsm que serão atualizadas durante a executção do script. 
        # Após a atualização essas planilhas serão protegidas novamente
        # https://docs.xlwings.org/en/stable/api.html#xlwings.Book.macro
        # ===================================================================================================
        desproteger = app.macro('Desproteger')
        desproteger()
        
        sheet = wb.sheets['Normais_Dados']
        last_row = sheet.range(1,1).end('down').row
        if 0 < last_row < 1048576:
            last_row = last_row + 1
            sheet.range("A{row}".format(row=last_row)).options(index=False,header=False).value = normal_df
        else:   
            sheet.range("A2").options(index=False,header=False).value = normal_df     
        sheet = wb.sheets['DayTrade_Dados']
        last_row2 = sheet.range(1,1).end('down').row
        if 0 < last_row2 < 1048576:
            last_row2 = last_row2 + 1
            sheet.range("A{row}".format(row=last_row2)).options(index=False,header=False).value = daytrade_df
        else:
            sheet.range("A2").options(index=False,header=False).value = daytrade_df       
        proteger = app.macro('Proteger')
        proteger()
        wb.save()
        wb.close()
        app.quit()        
    except FileNotFoundError:
        with pd.ExcelWriter(current_path+cpf+"/Completo.xlsx") as writer:
            note_df.to_excel(writer, sheet_name="Completo", index = False,header= True)
            normal_df.to_excel(writer, sheet_name="Normal", index = False,header= True)
            daytrade_df.to_excel(writer, sheet_name="DayTrade", index = False,header= True)
            taxas_df.to_excel(writer, sheet_name="Taxas", index = False,header= True)
    except PermissionError:
        print_erro()
        print('\n','O arquivo',current_path+cpf+"/Completo.xlsx",'está aberto, por favor feche-o e tente novamente')
        log += 'ERRO:\n' 
        log += ' - O arquivo'+current_path+cpf+"/Completo.xlsx"+' está aberto, por favor feche-o e tente novamente\n'
        '''Criar um controle para não mover os arquivos na rotina move_saida'''


# ===================================================================================================
# Cria o caminho completo de pastas/subpastas para mover os arquivos já processados.
# Os arquivos serão movidos de ./Entrada para a pasta ./Saida/CPF/Corretora/Ano
# ===================================================================================================
def move_saida(cpf,corretora,ano,mes,item):
    from os import makedirs
    log = ''
    destino_path = './Saida/'
    destino_prefix = cpf+'/'+corretora+'/'+ano+'/'+mes
    path_destino = join(destino_path, destino_prefix)   
    if exists(path_destino):
        print(f'Movendo o arquivo {basename(item)} para a pasta {path_destino}','\n')
        log = 'Movendo o arquivo "'+basename(item)+'" para a pasta '+path_destino+'.\n'
    if not exists('./Saida/'+cpf):
        makedirs(path_destino)
        print(f'Movendo o arquivo {basename(item)} para a pasta {path_destino}','\n')
        log = 'Movendo o arquivo "'+basename(item)+'" para a pasta '+path_destino+'.\n'
    elif not exists(path_destino):
        makedirs(path_destino)
        print(f'Movendo o arquivo {basename(item)} para a pasta {path_destino}','\n')
        log = 'Movendo o arquivo "'+basename(item)+'" para a pasta '+path_destino+'.\n'        
    try:
        shutil.move(item, join(path_destino, basename(item)))
        #print('Movendo o arquivo "{}" para a pasta "{}"'.format(basename(item), (path_destino)),'\n')
    except PermissionError:
        #print('\n')
        print_atencao()
        print(f'O arquivo {basename(item)} está aberto e não pode ser movido para a pasta {path_destino}','\n')
        print('Feche-o e mova-o manualmente','\n')
        log += 'ATENÇÃO:\n'
        log += ' - O arquivo "'+basename(item)+'" está aberto e não pode ser movido para a pasta "'+path_destino+'".\n'
        log += ' - Feche-o e mova-o manualmente.\n'
    return log

# ===================================================================================================
# Cria um arquivo de LOG para armazenar os dados do processamento
# ===================================================================================================
def log_processamento(current_path,cpf,log):
    nome_log = 'log_'+cpf+'.txt'
    ordem = list(dict.fromkeys(log))
    log_sem_repeticao = ''
    
    for n in range(0,len(ordem)):
            log_sem_repeticao += ordem[n]

    try:
        arquivo = open(current_path+cpf+'/'+nome_log,'a')
    except FileNotFoundError:
        arquivo = open(current_path+cpf+'/'+nome_log, 'w+')
    
    arquivo.write(log_sem_repeticao + '\n')
    arquivo.close()

# ===================================================================================================
# "Desenhos" para direcionar a atenção ou erro
# ===================================================================================================
def print_atencao():
    # Na tela sai legal...aqui está distorcido, provavelmente em função da largura dos caracteres
    # teria que ter uma fonte com largura fixa
    print("┌─────────────────┐")
    print("│  A T E N Ç Ã O  │")
    print("└─────────────────┘")

def print_erro():
    # Na tela sai legal...aqui está distorcido, provavelmente em função da largura dos caracteres
    # teria que ter uma fonte com largura fixa
    print("┌───────────┐")
    print("│  E R R O  │")
    print("└───────────┘")